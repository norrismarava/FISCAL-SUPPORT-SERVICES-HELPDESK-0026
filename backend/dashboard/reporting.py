from datetime import timedelta
from decimal import Decimal
from io import StringIO
import csv
from io import BytesIO

from django.conf import settings
from django.db.models import Count, Sum, Q
from django.utils import timezone
from django.utils.dateparse import parse_date, parse_datetime
from django.core.mail import send_mail

from callogs.models import CallLog
from tickets.models import SupportTicket
from users.models import User


def period_bounds(interval, now=None):
    now = now or timezone.now()
    if interval == 'hourly':
        start = now - timedelta(hours=1)
    elif interval == 'daily':
        start = now - timedelta(days=1)
    elif interval == 'monthly':
        start = now - timedelta(days=30)
    elif interval == 'quarterly':
        start = now - timedelta(days=90)
    elif interval == 'yearly':
        start = now - timedelta(days=365)
    else:
        start = now - timedelta(days=1)
    return start, now


def _coerce_list(value):
    if value is None:
        return []
    if isinstance(value, (list, tuple, set)):
        return [item for item in value if str(item).strip()]
    if isinstance(value, str):
        return [part.strip() for part in value.split(',') if part.strip()]
    return [value]


def _resolve_date_range(interval='daily', filters=None):
    filters = filters or {}
    start, end = period_bounds(interval)

    raw_start = filters.get('date_from') or filters.get('start')
    raw_end = filters.get('date_to') or filters.get('end')

    if raw_start:
        parsed = parse_datetime(raw_start)
        if parsed is None:
            d = parse_date(raw_start)
            if d is not None:
                parsed = timezone.make_aware(timezone.datetime.combine(d, timezone.datetime.min.time()))
        if parsed is not None:
            start = parsed

    if raw_end:
        parsed = parse_datetime(raw_end)
        if parsed is None:
            d = parse_date(raw_end)
            if d is not None:
                parsed = timezone.make_aware(timezone.datetime.combine(d, timezone.datetime.max.time()))
        if parsed is not None:
            end = parsed

    return start, end


def _apply_report_filters(jobs, tickets, filters=None):
    filters = filters or {}

    technician_ids = _coerce_list(filters.get('technician_ids') or filters.get('technician_id'))
    service_type_ids = _coerce_list(filters.get('service_type_ids') or filters.get('service_type_id'))
    fault_types = _coerce_list(filters.get('fault_types') or filters.get('fault_type'))
    ticket_statuses = _coerce_list(filters.get('ticket_statuses') or filters.get('ticket_status'))
    job_statuses = _coerce_list(filters.get('job_statuses') or filters.get('job_status'))
    regions = _coerce_list(filters.get('regions') or filters.get('region'))
    priorities = _coerce_list(filters.get('priorities') or filters.get('priority'))
    created_by_ids = _coerce_list(filters.get('created_by_ids') or filters.get('created_by_id'))

    if technician_ids:
        jobs = jobs.filter(assigned_technician_id__in=technician_ids)
        tickets = tickets.filter(assigned_to_id__in=technician_ids)
    if service_type_ids:
        tickets = tickets.filter(service_type_id__in=service_type_ids)
    if fault_types:
        jobs = jobs.filter(fault_type__in=fault_types)
    if ticket_statuses:
        tickets = tickets.filter(status__in=ticket_statuses)
    if job_statuses:
        jobs = jobs.filter(status__in=job_statuses)
    if regions:
        tickets = tickets.filter(region__in=regions)
    if priorities:
        tickets = tickets.filter(priority__in=priorities)
    if created_by_ids:
        jobs = jobs.filter(created_by_id__in=created_by_ids)

    return jobs, tickets


def get_default_report_recipients():
    recipients = set(
        User.objects.filter(role='accounts', is_active=True).exclude(email='').values_list('email', flat=True)
    )
    for key in ['REPORT_MR_DANIEL_EMAIL', 'REPORT_MR_TAPIWA_EMAIL']:
        value = getattr(settings, key, '')
        if value:
            recipients.add(value)

    extra = getattr(settings, 'REPORT_STATIC_RECIPIENTS', []) or []
    recipients.update([email for email in extra if email])
    return sorted(recipients)


def build_report_payload(interval='daily', filters=None):
    filters = filters or {}
    start, end = _resolve_date_range(interval=interval, filters=filters)
    jobs = CallLog.objects.filter(created_at__gte=start, created_at__lte=end)
    tickets = SupportTicket.objects.filter(created_at__gte=start, created_at__lte=end)
    jobs, tickets = _apply_report_filters(jobs, tickets, filters=filters)

    completed_jobs = jobs.filter(status='complete')
    solved_tickets = tickets.filter(status='solved')

    job_breakdown = list(
        jobs.values('fault_type').annotate(
            count=Count('id'),
            revenue=Sum('amount_charged'),
        ).order_by('-count')
    )
    ticket_breakdown = list(
        tickets.values('status').annotate(count=Count('id')).order_by('-count')
    )
    currency_breakdown = list(
        completed_jobs.values('currency').annotate(total=Sum('amount_charged')).order_by('currency')
    )

    # Resolution analytics
    resolved_job_hours = []
    for job in completed_jobs.exclude(completed_at__isnull=True):
        if job.completed_at and job.created_at:
            resolved_job_hours.append((job.completed_at - job.created_at).total_seconds() / 3600.0)
    avg_job_resolution_hours = round(sum(resolved_job_hours) / len(resolved_job_hours), 2) if resolved_job_hours else 0.0

    resolved_ticket_hours = []
    for ticket in solved_tickets.exclude(solved_at__isnull=True):
        if ticket.solved_at and ticket.created_at:
            resolved_ticket_hours.append((ticket.solved_at - ticket.created_at).total_seconds() / 3600.0)
    avg_ticket_resolution_hours = round(sum(resolved_ticket_hours) / len(resolved_ticket_hours), 2) if resolved_ticket_hours else 0.0

    pending_jobs_count = jobs.filter(status__in=['pending', 'assigned', 'in_progress']).count()
    pending_tickets_count = tickets.filter(status__in=['pending', 'open', 'reopened', 'unassigned']).count()
    escalated_tickets_count = tickets.filter(
        Q(priority__in=['high', 'urgent']) | Q(status='reopened')
    ).count()

    total_discount = float(completed_jobs.aggregate(total=Sum('discount_amount'))['total'] or 0)
    estimated_revenue = float(completed_jobs.aggregate(total=Sum('amount_charged'))['total'] or 0)
    estimated_cost_rate = float(getattr(settings, 'REPORT_ESTIMATED_COST_RATE', 0.0) or 0.0)
    estimated_costs = round(estimated_revenue * estimated_cost_rate, 2)

    rated_tickets = tickets.filter(csat_score__isnull=False)
    csat_avg = rated_tickets.aggregate(avg=Sum('csat_score'))['avg'] or 0
    csat_count = rated_tickets.count()
    customer_satisfaction_avg = round((float(csat_avg) / csat_count), 2) if csat_count else None

    payload = {
        'interval': interval,
        'start': start,
        'end': end,
        'filters': filters,
        'jobs_total': jobs.count(),
        'jobs_completed': completed_jobs.count(),
        'tickets_total': tickets.count(),
        'estimated_revenue': estimated_revenue,
        'estimated_costs': estimated_costs,
        'total_discount': total_discount,
        'pending_jobs': pending_jobs_count,
        'pending_tickets': pending_tickets_count,
        'escalated_tickets': escalated_tickets_count,
        'avg_job_resolution_hours': avg_job_resolution_hours,
        'avg_ticket_resolution_hours': avg_ticket_resolution_hours,
        'customer_satisfaction_avg': customer_satisfaction_avg,
        'customer_satisfaction_count': csat_count,
        'job_breakdown': job_breakdown,
        'ticket_breakdown': ticket_breakdown,
        'currency_breakdown': currency_breakdown,
    }
    return payload


def report_as_text(payload, include_fields=None):
    include_fields = include_fields or []
    include_all = not include_fields

    lines = [
        f"Helpdesk Summary Report ({payload['interval'].upper()})",
        f"Period: {payload['start']} to {payload['end']}",
        '',
        f"Total Jobs Logged: {payload['jobs_total']}",
        f"Completed Jobs: {payload['jobs_completed']}",
        f"Total Tickets Logged: {payload['tickets_total']}",
        f"Estimated Revenue: {payload['estimated_revenue']}",
        f"Estimated Costs: {payload.get('estimated_costs', 0)}",
        f"Pending Jobs: {payload.get('pending_jobs', 0)}",
        f"Pending Tickets: {payload.get('pending_tickets', 0)}",
        f"Escalated Tickets: {payload.get('escalated_tickets', 0)}",
        f"Average Job Resolution (hours): {payload.get('avg_job_resolution_hours', 0)}",
        f"Average Ticket Resolution (hours): {payload.get('avg_ticket_resolution_hours', 0)}",
        f"Customer Satisfaction Avg: {payload.get('customer_satisfaction_avg', 'N/A')}",
        f"Customer Satisfaction Responses: {payload.get('customer_satisfaction_count', 0)}",
        '',
    ]

    if include_all or 'job_breakdown' in include_fields:
        lines.append('Job Breakdown by Service Type/Fault Type:')
        for row in payload['job_breakdown']:
            revenue = row['revenue'] or 0
            lines.append(f"- {row['fault_type']}: {row['count']} job(s), revenue {revenue}")
        lines.append('')

    if include_all or 'ticket_breakdown' in include_fields:
        lines.append('Ticket Breakdown by Status:')
        for row in payload['ticket_breakdown']:
            lines.append(f"- {row['status']}: {row['count']} ticket(s)")
        lines.append('')

    if include_all or 'currency_breakdown' in include_fields:
        lines.append('Revenue by Currency:')
        for row in payload['currency_breakdown']:
            lines.append(f"- {row['currency']}: {row['total'] or 0}")
        lines.append('')

    return '\n'.join(lines)


def send_report(interval='daily', recipients=None, include_fields=None, filters=None):
    payload = build_report_payload(interval=interval, filters=filters)
    report_text = report_as_text(payload, include_fields=include_fields)
    recipients = recipients or get_default_report_recipients()

    if recipients:
        send_mail(
            subject=f'FSS Helpdesk {interval.title()} Summary Report',
            message=report_text,
            from_email=settings.DEFAULT_FROM_EMAIL,
            recipient_list=recipients,
            fail_silently=True,
        )

    return payload, recipients


def _decimal_or_blank(value):
    if value is None or value == '':
        return ''
    return Decimal(value)


def _hourly_rate(amount_charged, billed_hours):
    try:
        if billed_hours is None:
            return ''
        if isinstance(billed_hours, str) and '%' in billed_hours:
            return ''
        hours = Decimal(str(billed_hours))
        if hours <= 0:
            return ''
        return (Decimal(str(amount_charged or 0)) / hours).quantize(Decimal('0.01'))
    except Exception:
        return ''


def report_as_csv(payload, include_fields=None):
    """
    Build an Excel-friendly CSV report with summary and detailed sections.
    """
    include_fields = include_fields or []
    include_all = not include_fields

    output = StringIO()
    writer = csv.writer(output)

    writer.writerow(['FSS Helpdesk Summary Report'])
    writer.writerow(['Interval', (payload.get('interval') or '').upper()])
    writer.writerow(['Period Start', payload.get('start')])
    writer.writerow(['Period End', payload.get('end')])
    writer.writerow(['Applied Filters', payload.get('filters') or {}])
    writer.writerow([])

    writer.writerow(['Summary'])
    writer.writerow(['Metric', 'Value'])
    writer.writerow(['Total Jobs Logged', payload.get('jobs_total', 0)])
    writer.writerow(['Completed Jobs', payload.get('jobs_completed', 0)])
    writer.writerow(['Total Tickets Logged', payload.get('tickets_total', 0)])
    writer.writerow(['Estimated Revenue', payload.get('estimated_revenue', 0)])
    writer.writerow(['Estimated Costs', payload.get('estimated_costs', 0)])
    writer.writerow(['Total Discounts', payload.get('total_discount', 0)])
    writer.writerow(['Pending Jobs', payload.get('pending_jobs', 0)])
    writer.writerow(['Pending Tickets', payload.get('pending_tickets', 0)])
    writer.writerow(['Escalated Tickets', payload.get('escalated_tickets', 0)])
    writer.writerow(['Avg Job Resolution (Hours)', payload.get('avg_job_resolution_hours', 0)])
    writer.writerow(['Avg Ticket Resolution (Hours)', payload.get('avg_ticket_resolution_hours', 0)])
    writer.writerow(['Customer Satisfaction Avg', payload.get('customer_satisfaction_avg', 'N/A')])
    writer.writerow(['Customer Satisfaction Responses', payload.get('customer_satisfaction_count', 0)])
    writer.writerow([])

    # Main detailed jobs table.
    jobs_qs = CallLog.objects.filter(created_at__gte=payload.get('start'), created_at__lte=payload.get('end'))
    jobs_qs, _ = _apply_report_filters(jobs_qs, SupportTicket.objects.none(), filters=payload.get('filters') or {})
    jobs = (
        jobs_qs
        .select_related('assigned_technician', 'created_by')
        .prefetch_related('engineer_comments')
        .order_by('-created_at')
    )
    writer.writerow(['Main Jobs Sheet'])
    writer.writerow([
        'Job ID',
        'Job Card',
        'Customer Name',
        'Customer Email',
        'Customer Phone',
        'Customer Address',
        'Company Name',
        'Fault Type',
        'Fault Description',
        'ZIMRA Reference',
        'Date Booked',
        'Date Resolved',
        'Time Start',
        'Time Finish',
        'Job Type',
        'Status',
        'Billed Hours',
        'Amount Charged',
        'Hourly Rate',
        'Currency',
        'Assigned Technician',
        'Approved By',
        'Engineer Comments',
        'Booked By',
        'Created At',
        'Updated At',
    ])
    for job in jobs:
        latest_comment = job.engineer_comments.last()
        writer.writerow([
            str(job.job_id) if job.job_id else '',
            job.job_number or '',
            job.customer_name or '',
            job.customer_email or '',
            job.customer_phone or '',
            job.customer_address or '',
            'N/A',
            job.get_fault_type_display() if job.fault_type else 'N/A',
            job.fault_description or job.get_fault_type_display() or '',
            job.zimra_reference or '',
            job.booking_date or '',
            job.resolution_date or '',
            job.time_start or '',
            job.time_finish or '',
            job.get_job_type_display() if job.job_type else '',
            job.get_status_display() if job.status else '',
            job.billed_hours or '',
            _decimal_or_blank(job.amount_charged),
            _hourly_rate(job.amount_charged, job.billed_hours),
            job.currency or '',
            job.assigned_technician.get_full_name() if job.assigned_technician else '',
            '',
            latest_comment.comment if latest_comment else (job.resolution_notes or ''),
            job.created_by.get_full_name() if job.created_by else '',
            job.created_at,
            job.updated_at,
        ])
    writer.writerow([])

    if include_all or 'job_breakdown' in include_fields:
        writer.writerow(['Job Breakdown'])
        writer.writerow(['Fault Type', 'Jobs Count', 'Revenue'])
        for row in payload.get('job_breakdown', []):
            writer.writerow([
                row.get('fault_type') or 'N/A',
                row.get('count') or 0,
                Decimal(row.get('revenue') or 0),
            ])
        writer.writerow([])

    if include_all or 'ticket_breakdown' in include_fields:
        writer.writerow(['Ticket Breakdown'])
        writer.writerow(['Status', 'Tickets Count'])
        for row in payload.get('ticket_breakdown', []):
            writer.writerow([row.get('status') or 'N/A', row.get('count') or 0])
        writer.writerow([])

    if include_all or 'currency_breakdown' in include_fields:
        writer.writerow(['Revenue By Currency'])
        writer.writerow(['Currency', 'Amount'])
        for row in payload.get('currency_breakdown', []):
            writer.writerow([row.get('currency') or 'N/A', Decimal(row.get('total') or 0)])
        writer.writerow([])

    return output.getvalue()


def report_as_pdf(payload, include_fields=None):
    """
    Build a lightweight PDF summary export.
    Requires `reportlab`.
    """
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.pdfgen import canvas
    except Exception as exc:
        raise RuntimeError('PDF export requires reportlab. Install with `pip install reportlab`.') from exc

    text = report_as_text(payload, include_fields=include_fields)
    lines = text.split('\n')

    # reportlab canvas works with binary buffer
    from io import BytesIO
    pdf_buffer = BytesIO()
    c = canvas.Canvas(pdf_buffer, pagesize=A4)
    width, height = A4
    y = height - 40
    c.setFont('Helvetica-Bold', 12)
    c.drawString(40, y, 'FSS Helpdesk Report')
    y -= 24
    c.setFont('Helvetica', 10)
    for line in lines:
        if y < 40:
            c.showPage()
            c.setFont('Helvetica', 10)
            y = height - 40
        c.drawString(40, y, str(line)[:130])
        y -= 14
    c.save()
    return pdf_buffer.getvalue()


def report_as_xlsx(payload, include_fields=None):
    """
    Build a styled Excel report with consistent column sizing and readable colors.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except Exception as exc:
        raise RuntimeError('XLSX export requires openpyxl. Install with `pip install openpyxl`.') from exc

    include_fields = include_fields or []
    include_all = not include_fields

    wb = Workbook()
    ws = wb.active
    ws.title = 'Main Monthly Report'

    # Palette
    dark_blue = PatternFill(fill_type='solid', fgColor='1F2A44')
    teal = PatternFill(fill_type='solid', fgColor='0D9488')
    light_header = PatternFill(fill_type='solid', fgColor='334155')
    white_font = Font(color='FFFFFF', bold=True)
    section_font = Font(color='FFFFFF', bold=True, size=12)
    normal_font = Font(color='0F172A', size=10)
    thin_border = Border(
        left=Side(style='thin', color='94A3B8'),
        right=Side(style='thin', color='94A3B8'),
        top=Side(style='thin', color='94A3B8'),
        bottom=Side(style='thin', color='94A3B8'),
    )

    current_row = 1

    # Title block
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=10)
    title_cell = ws.cell(row=current_row, column=1, value='FSS HELP DESK - MAIN MONTHLY REPORT')
    title_cell.fill = dark_blue
    title_cell.font = Font(color='FFFFFF', bold=True, size=14)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    current_row += 1

    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=10)
    subtitle = f"Interval: {(payload.get('interval') or '').upper()} | Period: {payload.get('start')} to {payload.get('end')}"
    subtitle_cell = ws.cell(row=current_row, column=1, value=subtitle)
    subtitle_cell.fill = teal
    subtitle_cell.font = white_font
    subtitle_cell.alignment = Alignment(horizontal='center', vertical='center')
    current_row += 2

    # Summary section
    ws.cell(row=current_row, column=1, value='SUMMARY').fill = dark_blue
    ws.cell(row=current_row, column=1).font = section_font
    ws.cell(row=current_row, column=1).alignment = Alignment(horizontal='left', vertical='center')
    current_row += 1

    summary_rows = [
        ('Total Jobs Logged', payload.get('jobs_total', 0)),
        ('Completed Jobs', payload.get('jobs_completed', 0)),
        ('Total Tickets Logged', payload.get('tickets_total', 0)),
        ('Estimated Revenue', payload.get('estimated_revenue', 0)),
        ('Estimated Costs', payload.get('estimated_costs', 0)),
        ('Total Discounts', payload.get('total_discount', 0)),
        ('Pending Jobs', payload.get('pending_jobs', 0)),
        ('Pending Tickets', payload.get('pending_tickets', 0)),
        ('Escalated Tickets', payload.get('escalated_tickets', 0)),
        ('Avg Job Resolution (Hours)', payload.get('avg_job_resolution_hours', 0)),
        ('Avg Ticket Resolution (Hours)', payload.get('avg_ticket_resolution_hours', 0)),
        ('Customer Satisfaction Avg', payload.get('customer_satisfaction_avg', 'N/A')),
        ('Customer Satisfaction Responses', payload.get('customer_satisfaction_count', 0)),
    ]
    for metric, value in summary_rows:
        ws.cell(row=current_row, column=1, value=metric)
        ws.cell(row=current_row, column=2, value=value)
        ws.cell(row=current_row, column=1).font = normal_font
        ws.cell(row=current_row, column=2).font = normal_font
        ws.cell(row=current_row, column=1).border = thin_border
        ws.cell(row=current_row, column=2).border = thin_border
        current_row += 1

    current_row += 1

    # Main jobs table
    ws.cell(row=current_row, column=1, value='MAIN JOBS SHEET').fill = dark_blue
    ws.cell(row=current_row, column=1).font = section_font
    current_row += 1

    headers = [
        'Job ID',
        'Job Card',
        'Customer Name',
        'Customer Email',
        'Customer Phone',
        'Customer Address',
        'Company Name',
        'Fault Type',
        'Fault Description',
        'ZIMRA Reference',
        'Date Booked',
        'Date Resolved',
        'Time Start',
        'Time Finish',
        'Job Type',
        'Status',
        'Billed Hours',
        'Amount Charged',
        'Hourly Rate',
        'Currency',
        'Assigned Technician',
        'Approved By',
        'Engineer Comments',
        'Booked By',
        'Created At',
        'Updated At',
    ]

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=current_row, column=col_idx, value=header)
        cell.fill = light_header
        cell.font = white_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
    current_row += 1

    jobs_qs = CallLog.objects.filter(created_at__gte=payload.get('start'), created_at__lte=payload.get('end'))
    jobs_qs, _ = _apply_report_filters(jobs_qs, SupportTicket.objects.none(), filters=payload.get('filters') or {})
    jobs = (
        jobs_qs
        .select_related('assigned_technician', 'created_by')
        .prefetch_related('engineer_comments')
        .order_by('-created_at')
    )

    for job in jobs:
        latest_comment = job.engineer_comments.last()
        row_values = [
            str(job.job_id) if job.job_id else 'N/A',
            job.job_number or 'N/A',
            job.customer_name or 'N/A',
            job.customer_email or 'N/A',
            job.customer_phone or 'N/A',
            job.customer_address or 'N/A',
            job.customer_name or 'N/A',
            job.get_fault_type_display() if job.fault_type else 'N/A',
            job.fault_description or job.get_fault_type_display() or 'N/A',
            job.zimra_reference or 'N/A',
            str(job.booking_date or 'N/A'),
            str(job.resolution_date or 'N/A'),
            str(job.time_start or 'N/A'),
            str(job.time_finish or 'N/A'),
            job.get_job_type_display() if job.job_type else 'N/A',
            job.get_status_display() if job.status else 'N/A',
            job.billed_hours or 'N/A',
            float(job.amount_charged or 0),
            float(_hourly_rate(job.amount_charged, job.billed_hours) or 0) if _hourly_rate(job.amount_charged, job.billed_hours) != '' else '',
            job.currency or 'N/A',
            job.assigned_technician.get_full_name() if job.assigned_technician else 'Unassigned',
            '',
            latest_comment.comment if latest_comment else (job.resolution_notes or 'N/A'),
            job.created_by.get_full_name() if job.created_by else 'N/A',
            str(job.created_at or ''),
            str(job.updated_at or ''),
        ]
        for col_idx, value in enumerate(row_values, start=1):
            cell = ws.cell(row=current_row, column=col_idx, value=value)
            cell.font = normal_font
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            cell.border = thin_border
        current_row += 1

    # Keep columns and rows neat/consistent
    uniform_width = 20
    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = uniform_width

    for row_idx in range(1, current_row + 1):
        ws.row_dimensions[row_idx].height = 22

    # Optional extra sheets
    if include_all or 'job_breakdown' in include_fields:
        ws_jobs = wb.create_sheet('Job Breakdown')
        ws_jobs.append(['Fault Type', 'Jobs Count', 'Revenue'])
        for row in payload.get('job_breakdown', []):
            ws_jobs.append([row.get('fault_type') or 'N/A', row.get('count') or 0, float(row.get('revenue') or 0)])
        for c in range(1, 4):
            head = ws_jobs.cell(row=1, column=c)
            head.fill = light_header
            head.font = white_font
            head.border = thin_border
            ws_jobs.column_dimensions[get_column_letter(c)].width = 24

    if include_all or 'ticket_breakdown' in include_fields:
        ws_tickets = wb.create_sheet('Ticket Breakdown')
        ws_tickets.append(['Status', 'Tickets Count'])
        for row in payload.get('ticket_breakdown', []):
            ws_tickets.append([row.get('status') or 'N/A', row.get('count') or 0])
        for c in range(1, 3):
            head = ws_tickets.cell(row=1, column=c)
            head.fill = light_header
            head.font = white_font
            head.border = thin_border
            ws_tickets.column_dimensions[get_column_letter(c)].width = 24

    if include_all or 'currency_breakdown' in include_fields:
        ws_currency = wb.create_sheet('Revenue by Currency')
        ws_currency.append(['Currency', 'Amount'])
        for row in payload.get('currency_breakdown', []):
            ws_currency.append([row.get('currency') or 'N/A', float(row.get('total') or 0)])
        for c in range(1, 3):
            head = ws_currency.cell(row=1, column=c)
            head.fill = light_header
            head.font = white_font
            head.border = thin_border
            ws_currency.column_dimensions[get_column_letter(c)].width = 24

    stream = BytesIO()
    wb.save(stream)
    return stream.getvalue()
